#!/usr/bin/env python3
"""
DreamFinder Store Data Converter
=================================
Reads a completed onboarding spreadsheet (.xlsx) and generates:
  1. CSS variable block for store branding
  2. JavaScript MATTRESSES object
  3. JavaScript ACCESSORIES array
  4. Branding text replacements list
  5. (Optional) A full store-specific index.html

Usage:
    python convert_store_data.py <spreadsheet.xlsx> [options]

Common flags:
    --image-base-url URL       Public URL prefix where images will be served
    --source-images PATH       Auto-convert retailer images (mattresses/, accessories/) to optimized WebP
    --output-dir PATH          Where to write the build output (default: .)
    --output-html              Generate a complete index.html
    --skip-image-conversion    Skip the WebP conversion step

Examples:
    # Just emit the JSON / CSS / JS without touching images
    python convert_store_data.py "Store_Data.xlsx" --image-base-url "https://acme.github.io/DreamFinder"

    # Full build with image conversion (recommended for new retailer onboarding)
    python convert_store_data.py "Store_Data.xlsx" \\
        --image-base-url "https://acme.github.io/DreamFinder" \\
        --source-images "./incoming/" \\
        --output-dir "../"
"""

import argparse
import glob
import json
import os
import re
import sys

import openpyxl


# ── Image conversion ──────────────────────────────────────────────────────
# Source images for mattresses and accessories are converted to optimized
# WebP at build time. Keeps customer-facing pages small (~100x reduction
# vs. the original PNG/JPG sources) without asking the retailer to pre-shrink.

WEBP_LONG_EDGE = 1000
WEBP_QUALITY = 82
CONVERTIBLE_EXTS = ('.png', '.jpg', '.jpeg', '.PNG', '.JPG', '.JPEG', '.webp', '.WEBP')


def convert_images_to_webp(src_dir, dst_dir, label='images'):
    """Convert every image in src_dir into optimized .webp in dst_dir.

    - Resizes so the long edge is at most WEBP_LONG_EDGE
    - Encodes as WebP at WEBP_QUALITY
    - Output filename is the source stem + '.webp' (lowercased)
    - Idempotent: re-converting the same source overwrites the .webp

    Returns a list of (src_basename, src_bytes, dst_bytes) tuples.
    Requires Pillow (`pip install Pillow`).
    """
    try:
        from PIL import Image
    except ImportError:
        print("[!] Pillow not installed — skipping image conversion.")
        print("    Install with: pip install Pillow")
        return []

    if not os.path.isdir(src_dir):
        print(f"[!] {label} source folder not found: {src_dir} — skipping conversion.")
        return []

    os.makedirs(dst_dir, exist_ok=True)
    results = []
    for src_path in sorted(glob.glob(os.path.join(src_dir, '*'))):
        if not src_path.endswith(CONVERTIBLE_EXTS):
            continue
        base = os.path.basename(src_path)
        stem, _ = os.path.splitext(base)
        dst_path = os.path.join(dst_dir, stem.lower() + '.webp')

        try:
            img = Image.open(src_path)
            # WebP doesn't need alpha for product photos; flatten on white.
            if img.mode in ('RGBA', 'LA'):
                bg = Image.new('RGB', img.size, (255, 255, 255))
                bg.paste(img, mask=img.split()[-1])
                img = bg
            elif img.mode != 'RGB':
                img = img.convert('RGB')

            w, h = img.size
            if max(w, h) > WEBP_LONG_EDGE:
                if w >= h:
                    new_w = WEBP_LONG_EDGE
                    new_h = int(h * WEBP_LONG_EDGE / w)
                else:
                    new_h = WEBP_LONG_EDGE
                    new_w = int(w * WEBP_LONG_EDGE / h)
                img = img.resize((new_w, new_h), Image.LANCZOS)

            img.save(dst_path, 'WEBP', quality=WEBP_QUALITY, method=6)
            results.append((base, os.path.getsize(src_path), os.path.getsize(dst_path)))
        except Exception as e:
            print(f"[!] Failed to convert {base}: {e}")

    if results:
        before = sum(r[1] for r in results)
        after = sum(r[2] for r in results)
        print(f"  {label}: converted {len(results)} images "
              f"({before/1024/1024:.1f} MB → {after/1024/1024:.1f} MB, "
              f"{100*(1-after/before):.1f}% smaller)")
    return results


def read_store_info(ws):
    """Read the Store Info tab. Returns dict of store config."""
    headers = [cell.value for cell in ws[1]]
    # Find the first non-example row (row 3 if example is row 2, or row 2 if no example)
    data_row = None
    for row in ws.iter_rows(min_row=2, max_row=10, values_only=False):
        vals = [cell.value for cell in row]
        # Skip rows that are empty or the Bel Furniture example
        if not any(vals):
            continue
        if vals[0] and "Bel Furniture" in str(vals[0]):
            continue
        data_row = vals
        break

    # Fallback: if no non-example row found, use row 2 (might be the example itself for testing)
    if data_row is None:
        data_row = [cell.value for cell in ws[2]]

    mapping = {
        "Store Name": "store_name",
        "Logo Line 1": "logo_line1",
        "Logo Line 2": "logo_line2",
        "Primary Color (hex)": "primary_color",
        "Primary Color Light (hex)": "primary_color_light",
        "Trust Signal Text": "trust_signal",
        "Badge Text": "badge_text",
        "Default Location": "default_location",
        "Default Discount %": "default_discount",
        "Email Sender Name": "email_sender",
        "Email Subject Line": "email_subject",
        "Contact Email": "contact_email",
        "Store Phone": "store_phone",
        "Store Address": "store_address",
        "Store Hours": "store_hours",
        "Footer Text": "footer_text",
    }

    info = {}
    for i, header in enumerate(headers):
        if header in mapping and i < len(data_row):
            info[mapping[header]] = data_row[i] if data_row[i] is not None else ""
    return info


def read_brands(ws, image_base_url):
    """Read the Brands tab. Returns list of brand dicts."""
    headers = [cell.value for cell in ws[1]]
    clean_headers = []
    for h in headers:
        if h:
            h = h.replace("*", "").strip().split("\n")[0].strip()
        clean_headers.append(h)

    brands = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        vals = dict(zip(clean_headers, row))
        brand_name = str(vals.get("Brand Name", "")).strip()
        logo_file = str(vals.get("Logo File Name", "") or "").strip()
        if brand_name:
            brands.append({
                "name": brand_name,
                "logoUrl": f"{image_base_url}/logos/{logo_file}" if logo_file else "",
                "logoFile": logo_file,
            })
    return brands


def read_mattresses(ws, image_base_url):
    """Read the Mattresses tab. Returns dict with gold/silver/bronze lists."""
    headers = [cell.value for cell in ws[1]]
    # Clean header names (remove * and whitespace)
    clean_headers = []
    for h in headers:
        if h:
            h = h.replace("*", "").strip().split("\n")[0].strip()
        clean_headers.append(h)

    tiers = {"gold": [], "silver": [], "bronze": []}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue

        vals = dict(zip(clean_headers, row))

        mattress_id = str(vals.get("ID", "")).strip()
        name = str(vals.get("Name", "")).strip()
        brand = str(vals.get("Brand", "")).strip()
        sub_brand = str(vals.get("Sub-Brand", "") or "").strip()
        tier = str(vals.get("Tier", "bronze")).strip().lower()
        firmness = int(vals.get("Firmness (1-10)", 5) or 5)

        # Parse comma-separated lists
        tags = [t.strip() for t in str(vals.get("Display Tags", "") or "").split(",") if t.strip()]
        features = [f.strip() for f in str(vals.get("Feature Keywords", "") or "").split(",") if f.strip()]

        # Build image URL. Source file may be .png/.jpg/etc., but auto-conversion
        # produces .webp — point the URL at the converted version.
        img_file = str(vals.get("Image File Name", "") or "").strip()
        if img_file:
            stem, _ = os.path.splitext(img_file)
            image_url = f"{image_base_url}/images/mattresses/{stem.lower()}.webp"
        else:
            image_url = ""

        # Build reasons dict
        reasons = {}
        reason_map = {
            "Why: Cooling": "cooling",
            "Why: Pressure Relief": "pressureRelief",
            "Why: Motion Isolation": "motionIsolation",
            "Why: Support": "support",
            "Why: Firmness Feel": None,  # key determined by firmness
            "Why: Durability": "durability",
            "Why: Default": "default",
        }

        for col_name, reason_key in reason_map.items():
            val = str(vals.get(col_name, "") or "").strip()
            if val:
                if reason_key is None:
                    # Determine firmness feel key
                    if firmness <= 3:
                        reason_key = "plush"
                    elif firmness <= 6:
                        reason_key = "medium"
                    else:
                        reason_key = "firm"
                reasons[reason_key] = val

        mattress = {
            "id": mattress_id,
            "name": name,
            "brand": brand,
            "subBrand": sub_brand,
            "firmness": firmness,
            "tags": tags,
            "features": features,
            "imageUrl": image_url,
            "reasons": reasons,
        }

        if tier in tiers:
            tiers[tier].append(mattress)
        else:
            tiers["bronze"].append(mattress)

    return tiers


def read_accessories(ws, image_base_url):
    """Read the Accessories tab. Returns list of accessory objects."""
    headers = [cell.value for cell in ws[1]]
    clean_headers = []
    for h in headers:
        if h:
            h = h.replace("*", "").strip().split("\n")[0].strip()
        clean_headers.append(h)

    accessories = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        vals = dict(zip(clean_headers, row))

        acc_id = str(vals.get("ID", "")).strip()
        name = str(vals.get("Name", "")).strip()
        category = str(vals.get("Category", "")).strip()
        sub_type = str(vals.get("Sub-Type", "") or "").strip()
        price = vals.get("Price", 0) or 0
        description = str(vals.get("Description", "") or "").strip()

        img_file = str(vals.get("Image File Name", "") or "").strip()
        if img_file:
            stem, _ = os.path.splitext(img_file)
            image = f"{image_base_url}/images/accessories/{stem.lower()}.webp"
        else:
            image = ""

        match_tags = [t.strip() for t in str(vals.get("Match Tags", "") or "").split(",") if t.strip()]

        # Build matchScores
        score_map = {
            "Score: Default": "default",
            "Score: Cooling": "cooling",
            "Score: Hot": "hot",
            "Score: Back Pain": "back_pain",
            "Score: Snoring": "snoring",
            "Score: Premium": "premium",
            "Score: Position Side": "position_side",
            "Score: Position Back": "position_back",
            "Score: Position Stomach": "position_stomach",
        }

        match_scores = {}
        for col_name, score_key in score_map.items():
            val = vals.get(col_name)
            if val is not None and val != "" and val != 0:
                match_scores[score_key] = int(val)

        accessory = {
            "id": acc_id,
            "name": name,
            "category": category,
            "price": price,
            "image": image,
            "description": description,
        }
        if sub_type:
            accessory["subType"] = sub_type
        accessory["matchTags"] = match_tags
        accessory["matchScores"] = match_scores

        accessories.append(accessory)

    return accessories


def format_mattress_js(mattress):
    """Format a single mattress as a JS object literal string."""
    tags_str = json.dumps(mattress["tags"])
    features_str = json.dumps(mattress["features"])
    reasons_parts = []
    for k, v in mattress["reasons"].items():
        # Escape single quotes in value
        v_escaped = v.replace("'", "\\'")
        reasons_parts.append(f"'{k}':'{v_escaped}'")
    reasons_str = "{ " + ", ".join(reasons_parts) + " }"

    return (
        f"        {{ id: '{mattress['id']}', name: '{mattress['name']}', "
        f"brand: '{mattress['brand']}', subBrand: '{mattress['subBrand']}', "
        f"firmness: {mattress['firmness']}, "
        f"tags: {tags_str}, "
        f"features: {features_str}, "
        f"imageUrl: '{mattress['imageUrl']}', "
        f"reasons: {reasons_str} }}"
    )


def format_accessory_js(acc):
    """Format a single accessory as a JS object literal string."""
    lines = []
    lines.append(f"      {{ id: '{acc['id']}', name: '{acc['name']}', "
                  f"category: '{acc['category']}', price: {acc['price']},")
    lines.append(f"        image: '{acc['image']}',")
    desc_escaped = acc['description'].replace("'", "\\'")
    lines.append(f"        description: '{desc_escaped}',")
    if acc.get("subType"):
        lines.append(f"        subType: '{acc['subType']}',")
    lines.append(f"        matchTags: {json.dumps(acc['matchTags'])}, "
                  f"matchScores: {json.dumps(acc['matchScores'])} }}")
    return "\n".join(lines)


def generate_css_block(info):
    """Generate CSS custom property overrides."""
    color = info.get("primary_color", "#8B1A1A")
    light = info.get("primary_color_light", "#a52525")
    # Derive glow from primary color
    r, g, b = int(color[1:3], 16), int(color[3:5], 16), int(color[5:7], 16)
    glow = f"rgba({r}, {g}, {b}, 0.15)"

    return f"""      /* Store Brand Colors */
      --store-primary: {color};        /* {info.get('store_name', '')} */
      --store-primary-light: {light};
      --store-primary-glow: {glow};"""


def generate_mattresses_js(tiers):
    """Generate the full MATTRESSES const."""
    lines = ["    const MATTRESSES = {"]

    for tier_name in ["gold", "silver", "bronze"]:
        mattresses = tiers.get(tier_name, [])
        lines.append(f"      {tier_name}: [")
        for m in mattresses:
            lines.append(format_mattress_js(m) + ",")
        lines.append("      ],")

    lines.append("    };")
    return "\n".join(lines)


def generate_accessories_js(accessories):
    """Generate the full ACCESSORIES const."""
    lines = ["    const ACCESSORIES = ["]
    for acc in accessories:
        lines.append(format_accessory_js(acc) + ",")
    lines.append("    ];")
    return "\n".join(lines)


def generate_footer_html(info, brands, image_base_url):
    """Generate the footer HTML with store name and brand logos."""
    footer_text = info.get("footer_text", f"Powered by DreamFinder")
    store = info.get("store_name", "")

    brand_tags = []
    for b in brands:
        logo_html = ""
        if b.get("logoUrl"):
            logo_html = f'<img src="{b["logoUrl"]}" alt="{b["name"]}" class="brand-logo" />'
        else:
            logo_html = f'<span class="brand-logo-placeholder"></span>'
        brand_tags.append(
            f'          <div class="brand-tag">\n'
            f'            {logo_html}\n'
            f'            <span class="brand-name">{b["name"]}</span>\n'
            f'          </div>'
        )

    brands_joined = '\n          <span class="brand-sep"></span>\n'.join(brand_tags)

    return f"""    <!-- Footer -->
    <footer class="footer">
      <div class="footer-left">
        {footer_text}
      </div>
      <div class="brand-logos">
        <span class="label">Our Brands</span>
        <div class="brands">
{brands_joined}
        </div>
      </div>
      <div style="width:100%; text-align:center; padding-top:0.5rem; font-size:0.45rem; color:rgba(248,246,241,0.25); line-height:1.5;">
        DreamFinder is a recommendation tool, not medical advice. Match scores are guidance only. Purchases subject to store policies.
        <a href="#" onclick="event.preventDefault();document.getElementById('privacyOverlay').classList.add('visible');" style="color:rgba(212,168,75,0.4); text-decoration:underline; pointer-events: auto;">Privacy & Terms</a>
      </div>
    </footer>"""


def generate_branding_replacements(info, brands):
    """Generate a list of text replacements needed in index.html."""
    store = info.get("store_name", "")
    brand_list = ", ".join(b["name"] for b in brands) if brands else "(no brands specified)"
    return f"""
BRANDING TEXT REPLACEMENTS
==========================
Search & replace these strings in index.html:

  "Bel Furniture"                          -> "{store}"
  "bel</span>"  (logo line 1)             -> "{info.get('logo_line1', '')}</span>"
  "furniture</span>" (logo line 2)        -> "{info.get('logo_line2', '')}</span>"
  "Proudly serving Texas families..."      -> "{info.get('trust_signal', '')}"
  "Made in Texas"                          -> "{info.get('badge_text', '')}"
  "Texas" (default location)               -> "{info.get('default_location', '')}"

Footer:
  - Replace footer text with: "{info.get('footer_text', '')}"
  - Replace brand logos with: {brand_list}
  - Brand logo images go in logos/ folder

Also update:
  - manifest.json: "DreamFinder -- {store}"
  - Code.gs: Replace "Bel Furniture" in email subject/sender/body
  - Google Apps Script: Deploy a new script instance for this store
"""


def main():
    parser = argparse.ArgumentParser(description="Convert DreamFinder onboarding spreadsheet to JS code")
    parser.add_argument("spreadsheet", help="Path to the .xlsx file")
    parser.add_argument("--image-base-url", default="https://example.github.io/DreamFinder",
                        help="Base URL where images will be hosted (no trailing slash)")
    parser.add_argument("--output-html", action="store_true",
                        help="Generate a complete index.html (requires ../index.html as template)")
    parser.add_argument("--output-dir", default=".",
                        help="Directory to write output files (default: current dir)")
    parser.add_argument("--source-images", default=None,
                        help="Path to retailer's submitted images folder (with mattresses/ and accessories/ subdirs). "
                             "If set, images are auto-converted to optimized WebP into <output-dir>/images/.")
    parser.add_argument("--skip-image-conversion", action="store_true",
                        help="Skip image WebP conversion (if you've already done it).")
    args = parser.parse_args()

    # Strip trailing slash from URL
    args.image_base_url = args.image_base_url.rstrip("/")

    # Auto-convert source images to WebP if a source folder was provided
    if args.source_images and not args.skip_image_conversion:
        print(f"Converting source images from {args.source_images}...")
        convert_images_to_webp(
            os.path.join(args.source_images, 'mattresses'),
            os.path.join(args.output_dir, 'images', 'mattresses'),
            label='mattresses'
        )
        convert_images_to_webp(
            os.path.join(args.source_images, 'accessories'),
            os.path.join(args.output_dir, 'images', 'accessories'),
            label='accessories'
        )

    print(f"Reading {args.spreadsheet}...")
    wb = openpyxl.load_workbook(args.spreadsheet, data_only=True)

    # Read all tabs
    store_info = read_store_info(wb["Store Info"])
    print(f"  Store: {store_info.get('store_name', '(unknown)')}")

    tiers = read_mattresses(wb["Mattresses"], args.image_base_url)
    total_m = sum(len(v) for v in tiers.values())
    print(f"  Mattresses: {total_m} (gold={len(tiers['gold'])}, silver={len(tiers['silver'])}, bronze={len(tiers['bronze'])})")

    accessories = read_accessories(wb["Accessories"], args.image_base_url)
    print(f"  Accessories: {len(accessories)}")

    brands = []
    if "Brands" in wb.sheetnames:
        brands = read_brands(wb["Brands"], args.image_base_url)
        print(f"  Brands: {len(brands)} ({', '.join(b['name'] for b in brands)})")

    # Generate outputs
    os.makedirs(args.output_dir, exist_ok=True)

    css_block = generate_css_block(store_info)
    mattresses_js = generate_mattresses_js(tiers)
    accessories_js = generate_accessories_js(accessories)
    footer_html = generate_footer_html(store_info, brands, args.image_base_url)
    replacements = generate_branding_replacements(store_info, brands)

    # Write individual output files
    store_slug = store_info.get("store_name", "store").lower().replace(" ", "-")

    css_path = os.path.join(args.output_dir, f"{store_slug}_css.txt")
    with open(css_path, "w", encoding="utf-8") as f:
        f.write(css_block)
    print(f"\n  CSS variables  -> {css_path}")

    matt_path = os.path.join(args.output_dir, f"{store_slug}_mattresses.js")
    with open(matt_path, "w", encoding="utf-8") as f:
        f.write(mattresses_js)
    print(f"  MATTRESSES     -> {matt_path}")

    acc_path = os.path.join(args.output_dir, f"{store_slug}_accessories.js")
    with open(acc_path, "w", encoding="utf-8") as f:
        f.write(accessories_js)
    print(f"  ACCESSORIES    -> {acc_path}")

    repl_path = os.path.join(args.output_dir, f"{store_slug}_replacements.txt")
    with open(repl_path, "w", encoding="utf-8") as f:
        f.write(replacements)
    print(f"  Replacements   -> {repl_path}")

    footer_path = os.path.join(args.output_dir, f"{store_slug}_footer.html")
    with open(footer_path, "w", encoding="utf-8") as f:
        f.write(footer_html)
    print(f"  Footer HTML    -> {footer_path}")

    # Optionally generate full HTML
    if args.output_html:
        template_path = os.path.join(os.path.dirname(__file__), "..", "index.html")
        if not os.path.exists(template_path):
            print(f"\n  ERROR: Template not found at {template_path}")
            sys.exit(1)

        with open(template_path, "r", encoding="utf-8") as f:
            html = f.read()

        # Replace CSS variables
        css_pattern = r"(--store-primary:)[^;]+(;.*?\n\s*--store-primary-light:)[^;]+(;.*?\n\s*--store-primary-glow:)[^;]+(;)"
        new_css = (
            f"\\1 {store_info.get('primary_color', '#8B1A1A')}\\2 "
            f"{store_info.get('primary_color_light', '#a52525')}\\3 "
        )
        r, g, b = (int(store_info['primary_color'][i:i+2], 16) for i in (1, 3, 5))
        glow = f"rgba({r}, {g}, {b}, 0.15)"
        new_css += f" {glow}\\4"
        html = re.sub(css_pattern, new_css, html, flags=re.DOTALL)

        # Replace MATTRESSES object
        matt_pattern = r"const MATTRESSES = \{.*?\};\s*\n"
        html = re.sub(matt_pattern, mattresses_js + "\n", html, flags=re.DOTALL)

        # Replace ACCESSORIES array
        acc_pattern = r"const ACCESSORIES = \[.*?\];\s*\n"
        html = re.sub(acc_pattern, accessories_js + "\n", html, flags=re.DOTALL)

        # Replace store name references
        html = html.replace("Bel Furniture", store_info.get("store_name", ""))
        html = html.replace(
            '<span class="logo-main">bel</span>',
            f'<span class="logo-main">{store_info.get("logo_line1", "")}</span>'
        )
        html = html.replace(
            '<span class="logo-sub">furniture</span>',
            f'<span class="logo-sub">{store_info.get("logo_line2", "")}</span>'
        )

        # Replace trust signal and badge
        html = html.replace(
            "Proudly serving Texas families for over 25 years",
            store_info.get("trust_signal", "")
        )
        html = html.replace("Made in Texas", store_info.get("badge_text", ""))

        # Replace footer
        footer_pattern = r"    <!-- Footer -->.*?    </footer>"
        html = re.sub(footer_pattern, footer_html, html, flags=re.DOTALL)

        # Replace image base URL
        html = html.replace(
            "https://beford782.github.io/DreamFinder",
            args.image_base_url
        )

        html_path = os.path.join(args.output_dir, f"{store_slug}_index.html")
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html)
        print(f"  Full HTML      -> {html_path}")

    print("\nDone!")


if __name__ == "__main__":
    main()
