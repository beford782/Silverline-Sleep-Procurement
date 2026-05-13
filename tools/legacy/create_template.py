#!/usr/bin/env python3
"""Generate the DreamFinder Store Onboarding spreadsheet template (.xlsx)."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = "../onboarding/DreamFinder_Onboarding_Template.xlsx"

# ── Styles ──────────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F3A5C", end_color="1F3A5C", fill_type="solid")
EXAMPLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
OPTIONAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="1F3A5C")
BODY_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, min_width=12, max_width=40):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        lengths = []
        for cell in col:
            if cell.value:
                lines = str(cell.value).split("\n")
                lengths.append(max(len(line) for line in lines))
        width = min(max(max(lengths, default=min_width), min_width), max_width)
        ws.column_dimensions[letter].width = width + 2


# ════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ── Tab 1: Store Info ──────────────────────────────────────────────────────
ws = wb.active
ws.title = "Store Info"

headers = [
    "Store Name",
    "Logo Line 1",
    "Logo Line 2",
    "Primary Color (hex)",
    "Primary Color Light (hex)",
    "Trust Signal Text",
    "Badge Text",
    "Default Location",
    "Default Discount %",
    "Email Sender Name",
    "Email Subject Line",
    "Contact Email",
    "Store Phone",
    "Store Address",
    "Store Hours",
    "Footer Text",
]
for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h)

# Example row (Bel Furniture)
example = [
    "Bel Furniture",
    "bel",
    "furniture",
    "#8B1A1A",
    "#a52525",
    "Proudly serving Texas families for over 25 years",
    "Made in Texas",
    "Texas",
    5,
    "Bel Furniture Sleep Team",
    "Your DreamFinder Results from Bel Furniture",
    "sleep@belfurniture.com",
    "(713) 555-0100",
    "12345 Main St, Houston, TX 77001",
    "Mon–Sat 10am–8pm · Sun 12–6pm",
    "Powered by DreamFinder · (c) 2026 Bel Furniture",
]
for i, val in enumerate(example, 1):
    cell = ws.cell(row=2, column=i, value=val)
    cell.fill = EXAMPLE_FILL
    cell.font = BODY_FONT
    cell.border = THIN_BORDER

# Row 3: blank for the store to fill
for i in range(1, len(headers) + 1):
    cell = ws.cell(row=3, column=i)
    cell.fill = REQUIRED_FILL
    cell.border = THIN_BORDER

style_header_row(ws, len(headers))
auto_width(ws)

# Note at top
ws.cell(row=5, column=1, value="YELLOW row = Bel Furniture example (delete before submitting). GREEN row = fill in your info here.").font = Font(italic=True, color="666666")

# ── Tab 2: Mattresses ─────────────────────────────────────────────────────
ws2 = wb.create_sheet("Mattresses")

matt_headers = [
    "ID *",
    "Name *",
    "Brand *",
    "Sub-Brand",
    "Tier *",
    "Firmness (1-10) *",
    "Display Tags\n(comma-sep)",
    "Feature Keywords\n(comma-sep, see ref tab)",
    "Image File Name *\n(e.g. athena.png)",
    "Why: Cooling",
    "Why: Pressure Relief",
    "Why: Motion Isolation",
    "Why: Support",
    "Why: Firmness Feel",
    "Why: Durability",
    "Why: Default *",
]
for i, h in enumerate(matt_headers, 1):
    ws2.cell(row=1, column=i, value=h)

# Example rows from Bel Furniture
bel_mattresses = [
    ["g4", "Athena", "Spring Air", "Last Mattress", "gold", 6,
     "Hybrid, Latex, Hand-Tufted",
     "medium, cooling, support, hybrid, responsive, latex, hand-tufting, lifetime-warranty",
     "athena.png",
     "Phase change technology — cool-to-the-touch fabrics",
     "Talalay latex for natural, durable pressure relief",
     "Pocketed coils minimize partner disturbance",
     "Hand-tufted QUANTUM ENCASED COIL UNIT with 1575 coils",
     "Perfect medium feel for most sleepers",
     "Hand-tufted, lifetime warranty",
     "Ultra premium hand-tufted hybrid with specialty foams and lifetime warranty"],
    ["g1", "Caitlin", "Spring Air", "", "gold", 2,
     "Hybrid, Copper",
     "plush, cooling, pressure-relief, hybrid, motion-isolation, copper",
     "caitlin.png",
     "Copper-infused foam conducts heat away",
     "Plush hybrid feel conforms to body curves",
     "Pocketed coils isolate movement",
     "Hybrid coil + foam construction",
     "Ultra plush feel for side sleepers",
     "",
     "Copper-infused plush hybrid with antimicrobial properties"],
    ["g20", "Bella", "Spring Air", "", "bronze", 3,
     "Innerspring, Back Supporter",
     "plush, comfort, support, innerspring, foam-encasement",
     "bella.png",
     "",
     "",
     "",
     "Back Supporter technology with foam encasement",
     "Plush comfort at an accessible price",
     "",
     "Comfortable innerspring with Back Supporter technology"],
]

for row_idx, row_data in enumerate(bel_mattresses, 2):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = EXAMPLE_FILL
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = WRAP

# Data validation: Tier dropdown
tier_dv = DataValidation(type="list", formula1='"gold,silver,bronze"', allow_blank=False)
tier_dv.error = "Tier must be gold, silver, or bronze"
tier_dv.errorTitle = "Invalid Tier"
ws2.add_data_validation(tier_dv)
tier_dv.add(f"E2:E100")

# Data validation: Firmness 1-10
firmness_dv = DataValidation(type="whole", operator="between", formula1=1, formula2=10)
firmness_dv.error = "Firmness must be a whole number from 1 to 10"
firmness_dv.errorTitle = "Invalid Firmness"
ws2.add_data_validation(firmness_dv)
firmness_dv.add(f"F2:F100")

style_header_row(ws2, len(matt_headers))
auto_width(ws2)

# ── Tab 3: Accessories ────────────────────────────────────────────────────
ws3 = wb.create_sheet("Accessories")

acc_headers = [
    "ID *",
    "Name *",
    "Category *",
    "Sub-Type",
    "Price *",
    "Description *",
    "Image File Name *",
    "Match Tags\n(comma-sep)",
    "Score: Default",
    "Score: Cooling",
    "Score: Hot",
    "Score: Back Pain",
    "Score: Snoring",
    "Score: Premium",
    "Score: Position Side",
    "Score: Position Back",
    "Score: Position Stomach",
]
for i, h in enumerate(acc_headers, 1):
    ws3.cell(row=1, column=i, value=h)

bel_accessories = [
    ["base-bt3000", "BT3000 Power Base", "Foundations & Support", "adjustable", 899,
     "Head & foot adjustment, wireless remote, USB ports, massage, LED lighting",
     "base-bt3000.jpg", "snoring, back_pain", "", "", "", 4, 4, 3, "", "", ""],
    ["pillow-activecool", "Active Cool Pillow", "Pillows", "", 160,
     "Advanced active cooling fabric with premium memory foam clusters",
     "pillow-activecool.webp", "hot_sleeper, cooling", "", 4, 4, "", "", "", 2, 2, ""],
    ["protector-activedry", "Active Dry Protector", "Protectors", "", 100,
     "Waterproof & breathable — wicks moisture, extends mattress life",
     "protector-activedry.jpg", "all", 3, 1, "", "", "", "", "", "", ""],
]

for row_idx, row_data in enumerate(bel_accessories, 2):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = EXAMPLE_FILL
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = WRAP

# Category dropdown
cat_dv = DataValidation(type="list", formula1='"Foundations & Support,Pillows,Protectors"', allow_blank=False)
cat_dv.error = "Must be: Foundations & Support, Pillows, or Protectors"
ws3.add_data_validation(cat_dv)
cat_dv.add("C2:C50")

# Sub-Type dropdown
sub_dv = DataValidation(type="list", formula1='"adjustable,foundation,low_profile,bunkie,"', allow_blank=True)
sub_dv.error = "Must be: adjustable, foundation, low_profile, bunkie, or blank"
ws3.add_data_validation(sub_dv)
sub_dv.add("D2:D50")

# Score validation (0-5)
score_dv = DataValidation(type="whole", operator="between", formula1=0, formula2=5)
score_dv.error = "Score must be 0-5"
ws3.add_data_validation(score_dv)
score_dv.add("I2:Q50")

style_header_row(ws3, len(acc_headers))
auto_width(ws3)

# ── Tab 4: Brands (Footer) ───────────────────────────────────────────────
ws_brands = wb.create_sheet("Brands")

brand_headers = [
    "Brand Name *",
    "Logo File Name *\n(e.g. restonic-logo.png)",
]
for i, h in enumerate(brand_headers, 1):
    ws_brands.cell(row=1, column=i, value=h)

bel_brands = [
    ["Restonic", "restonic-logo.png"],
    ["Spring Air", "springair-logo.png"],
    ["Bel-O-Pedic", "belopedic-logo.png"],
]

for row_idx, row_data in enumerate(bel_brands, 2):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws_brands.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = EXAMPLE_FILL
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = WRAP

style_header_row(ws_brands, len(brand_headers))
auto_width(ws_brands)

# ── Tab 5: Feature Keywords Reference ─────────────────────────────────────
ws4 = wb.create_sheet("Feature Keywords")

ws4.cell(row=1, column=1, value="Feature Keyword").font = HEADER_FONT
ws4.cell(row=1, column=1).fill = HEADER_FILL
ws4.cell(row=1, column=2, value="What It Means / When to Use").font = HEADER_FONT
ws4.cell(row=1, column=2).fill = HEADER_FILL
ws4.cell(row=1, column=3, value="Quiz Connection").font = HEADER_FONT
ws4.cell(row=1, column=3).fill = HEADER_FILL

keywords = [
    ("plush", "Very soft feel (firmness 1-3)", "Matches side sleepers, lighter body types"),
    ("medium", "Moderate feel (firmness 4-6)", "Most versatile — matches combo sleepers, average body types"),
    ("firm", "Hard feel (firmness 7-10)", "Matches back/stomach sleepers, heavier body types"),
    ("cooling", "Has cooling technology (gel, phase-change, copper)", "Matches hot sleepers, temperature complaints"),
    ("support", "Strong core support system", "Matches back pain, heavier body types"),
    ("hybrid", "Coils + foam construction", "General positive scoring — most versatile type"),
    ("innerspring", "Traditional coil construction", "Budget-friendly, good support"),
    ("latex", "Contains latex foam layers", "Natural feel, responsive, hypoallergenic"),
    ("responsive", "Quick response to movement (not slow-sinking)", "Combo sleepers, people who move a lot"),
    ("pressure-relief", "Relieves pressure points (hips, shoulders)", "Side sleepers, joint pain, shoulder issues"),
    ("motion-isolation", "Minimizes partner movement transfer", "Couples, light sleepers with partners"),
    ("hand-tufting", "Hand-tufted construction (premium)", "Premium/durability seekers"),
    ("copper", "Copper-infused materials", "Antimicrobial, cooling, wellness-focused"),
    ("gel-memory-foam", "Gel-infused memory foam", "Cooling + pressure relief"),
    ("gel-infused-foam", "Gel-infused foam (non-memory)", "Cooling + comfort"),
    ("serene-foam", "Serene foam comfort layer", "Comfort + cooling"),
    ("foam-encasement", "Foam-encased edge support", "Edge support, full sleep surface"),
    ("back-supporter-technology", "Back Supporter certified", "Back pain sufferers"),
    ("quality", "High-quality materials/construction", "Durability seekers"),
    ("comfort", "General comfort emphasis", "Broad appeal"),
    ("lifetime-warranty", "Comes with lifetime warranty", "Durability/value seekers"),
    ("anti-viral", "Antiviral material properties", "Health-conscious sleepers"),
    ("anti-bacterial", "Antibacterial material properties", "Health-conscious sleepers"),
    ("anti-microbial", "Antimicrobial material properties", "Health-conscious sleepers"),
    ("anti-inflammatory", "Anti-inflammatory material properties", "Health-conscious sleepers"),
    ("durability", "Built to last / premium construction", "Long-term investment seekers"),
]

for row_idx, (kw, desc, quiz) in enumerate(keywords, 2):
    ws4.cell(row=row_idx, column=1, value=kw).font = Font(name="Calibri", bold=True, size=11)
    ws4.cell(row=row_idx, column=2, value=desc).font = BODY_FONT
    ws4.cell(row=row_idx, column=3, value=quiz).font = BODY_FONT
    for col in range(1, 4):
        ws4.cell(row=row_idx, column=col).border = THIN_BORDER
        ws4.cell(row=row_idx, column=col).alignment = WRAP

auto_width(ws4, max_width=50)

# ── Tab 6: Instructions ───────────────────────────────────────────────────
ws5 = wb.create_sheet("Instructions")

instructions = [
    ("DreamFinder Store Onboarding — How to Fill Out This Spreadsheet", ""),
    ("", ""),
    ("STEP 1: STORE INFO TAB", ""),
    ("Fill in the GREEN row with your store's information.", ""),
    ("• Store Name — Your store's full name as you want it displayed (e.g., 'Texas Mattress Co')", ""),
    ("• Logo Line 1 / Line 2 — How your name appears in the app logo. Line 1 is large, Line 2 is smaller.", ""),
    ("• Primary Color — Your brand's main color as a hex code (e.g., #2B5797). Use a hex picker online.", ""),
    ("• Primary Color Light — A lighter shade of your brand color for hover effects.", ""),
    ("• Trust Signal Text — A tagline shown on the welcome screen (e.g., 'Family-owned since 1995').", ""),
    ("• Badge Text — A short badge label (e.g., 'Local Favorite', 'Family Owned').", ""),
    ("• Default Location — Your state or region, shown before GPS kicks in.", ""),
    ("• Default Discount % — The discount percentage for the quiz completion reward.", ""),
    ("• Email Sender Name — The 'From' name on result emails sent to customers.", ""),
    ("• Email Subject Line — The subject line of the customer results email.", ""),
    ("• Contact Email — Shown in the privacy section.", ""),
    ("• Store Phone — Phone number shown in the customer email and on the handoff screen.", ""),
    ("• Store Address — Street address shown in the customer email so they can find you.", ""),
    ("• Store Hours — Operating hours shown in the customer email (e.g., 'Mon–Sat 10am–8pm').", ""),
    ("", ""),
    ("STEP 2: MATTRESSES TAB", ""),
    ("Add one row per mattress you carry. Fields marked with * are required.", ""),
    ("• ID — A short unique code (e.g., 'athena', 'model-100'). Use lowercase, no spaces.", ""),
    ("• Name — The display name (e.g., 'Athena', 'Royal Cloud').", ""),
    ("• Brand — The manufacturer (e.g., 'Serta', 'Sealy').", ""),
    ("• Tier — gold (premium), silver (mid-range), or bronze (value). This controls display order.", ""),
    ("• Firmness — Rate 1 (ultra plush) to 10 (ultra firm).", ""),
    ("• Display Tags — Comma-separated labels shown on the card (e.g., 'Hybrid, Latex, Pillow-Top').", ""),
    ("• Feature Keywords — Comma-separated from the 'Feature Keywords' tab. These drive quiz matching.", ""),
    ("• Image File Name — Must exactly match the file you upload to the images folder.", ""),
    ("• Why columns — Short explanations of why this mattress is great for each quality.", ""),
    ("  'Why: Default' is required — it's the main selling point shown to every customer.", ""),
    ("", ""),
    ("STEP 3: ACCESSORIES TAB", ""),
    ("Add one row per accessory (bases, pillows, protectors).", ""),
    ("• Category must be: Foundations & Support, Pillows, or Protectors.", ""),
    ("• Sub-Type is for foundations only: adjustable, foundation, low_profile, or bunkie.", ""),
    ("• Match Scores (0-5) control how strongly this accessory is recommended based on quiz answers.", ""),
    ("  Leave blank or 0 if the score doesn't apply.", ""),
    ("", ""),
    ("STEP 4: BRANDS TAB", ""),
    ("Add one row per mattress brand you carry. These appear in the app footer.", ""),
    ("• Brand Name — The display name (e.g., 'Serta', 'Sealy').", ""),
    ("• Logo File Name — Must match a file in the logos/ folder (e.g., serta-logo.png).", ""),
    ("", ""),
    ("STEP 5: UPLOAD IMAGES TO THE SHARED DRIVE FOLDER", ""),
    ("Place your files in the shared Google Drive folder:", ""),
    ("  logos/        — Your store logo, PWA icons (192x192 and 512x512), and brand logos", ""),
    ("  mattresses/   — One image per mattress, named to match the ID column (e.g., athena.png)", ""),
    ("  accessories/  — One image per accessory, named to match the ID column (e.g., base-bt3000.jpg)", ""),
    ("", ""),
    ("IMAGE REQUIREMENTS:", ""),
    ("  Mattresses:  Up to 1000px on the long edge. WebP / JPG / PNG all accepted.", ""),
    ("               Anything you submit is auto-converted to WebP at build time, so", ""),
    ("               don't worry about size — give us the highest-quality source you have.", ""),
    ("  Accessories: Same — up to 1000px, any common format. Auto-converted to WebP.", ""),
    ("  Logo:        Min 400px wide, transparent background, PNG (NOT auto-converted).", ""),
    ("  PWA Icons:   Exactly 192×192 and 512×512 pixels, square, PNG (NOT auto-converted).", ""),
    ("", ""),
    ("TIPS:", ""),
    ("• The YELLOW example rows show Bel Furniture's data — delete them before submitting.", ""),
    ("• Use the Feature Keywords tab as a reference when filling in feature keywords.", ""),
    ("• You can add your OWN feature keywords if needed — just be consistent.", ""),
    ("• If you're unsure about match scores for accessories, leave them blank and we'll help.", ""),
]

for row_idx, (text, _) in enumerate(instructions, 1):
    cell = ws5.cell(row=row_idx, column=1, value=text)
    if text.startswith("STEP") or text.startswith("IMAGE") or text.startswith("TIPS"):
        cell.font = SECTION_FONT
    elif row_idx == 1:
        cell.font = Font(name="Calibri", bold=True, size=14, color="1F3A5C")
    else:
        cell.font = BODY_FONT

ws5.column_dimensions["A"].width = 100

# ── Save ───────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"Template saved to {OUTPUT}")
